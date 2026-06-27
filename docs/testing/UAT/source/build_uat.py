# Builds the comprehensive UAT workbook from the per-group JSON row files.
import json, glob, re, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

SCRATCH = r"C:/Users/Godfrey/AppData/Local/Temp/claude/d--My-Works-ERP-ERPCLEAN2/64a2e4ee-1bf4-4ea4-8643-22ae4bb82f42/scratchpad"
OUT = r"d:/My_Works/ERP/ERPCLEAN2/docs/testing/UAT/ERP-UAT-Test-Plan.xlsx"

rows = []
for f in sorted(glob.glob(SCRATCH + "/uat-rows-*.json")):
    rows.extend(json.load(open(f, encoding="utf-8-sig")))

# ---- desired module/sheet order ----
ORDER = [
    "IAM & Access","RBAC","Platform – Documents","Platform – Notifications",
    "Platform – Approvals","Platform – Audit","Conventions",
    "Masters – Parties","Masters – Catalog","Sales – Order to Cash","Sales – Pricing","POS",
    "Procurement (P2P)","Inventory","Manufacturing",
    "General Ledger","Accounts Receivable","Accounts Payable","Cash & Bank",
    "Tax (VAT/WHT)","FX & Multicurrency","Reporting & BI","Fixed Assets","Budgeting",
    "HR & Payroll","CRM","Projects",
]
mods = {}
for r in rows:
    mods.setdefault(r.get("module","Other"), []).append(r)
ordered_mods = [m for m in ORDER if m in mods] + [m for m in mods if m not in ORDER]

# ---- styling ----
NAVY = "1F3864"; BLUE = "2E5496"; LIGHT = "D9E1F2"; GREY = "F2F2F2"
HEAD_FILL = PatternFill("solid", fgColor=BLUE)
TITLE_FILL = PatternFill("solid", fgColor=NAVY)
SUB_FILL = PatternFill("solid", fgColor=LIGHT)
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
CTR = Alignment(horizontal="center", vertical="top")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = [
    ("ID",10),("Area",18),("Test Scenario",40),("Type",11),("Priority",9),
    ("Preconditions",34),("Steps",46),("Test Data",26),("Expected Result",42),
    ("Route",26),("Permission",22),("Actual Result",30),("Status",12),
    ("Severity",11),("Tester",14),("Run Date",12),("Comments",30),
]
STATUS_COL = 13  # 'Status' (M)
sheetname_of = {}

def safe(name):
    n = re.sub(r'[\\/?*\[\]:]', '-', name)[:31]
    base = n; i = 1
    while n in sheetname_of.values():
        n = (base[:28] + f"~{i}"); i += 1
    return n

wb = Workbook()

# ============ COVER ============
cov = wb.active; cov.title = "Cover & Instructions"
cov.sheet_view.showGridLines = False
cov.column_dimensions["A"].width = 3
cov.column_dimensions["B"].width = 26
cov.column_dimensions["C"].width = 100
def cov_row(r, label, val, label_bold=True):
    cov.cell(r,2,label).font = BOLD if label_bold else Font()
    c = cov.cell(r,3,val); c.alignment = WRAP; return c
cov.merge_cells("B2:C2")
t = cov.cell(2,2,"ERP — User Acceptance Test (UAT) Plan"); t.font = Font(size=18,bold=True,color="FFFFFF"); t.fill=TITLE_FILL; t.alignment=Alignment(horizontal="center",vertical="center")
cov.row_dimensions[2].height = 30
cov_row(4,"Purpose","End-to-end UAT for the ERP, organised one tab per module. Each row is a manually-executable test case for the QA team. Fill in Actual Result, Status, Severity, Tester and Run Date as you go; the Summary tab tallies results live.")
cov_row(5,"Total test cases", str(len(rows)))
cov_row(6,"Modules / tabs", str(len(ordered_mods)))
cov_row(7,"Generated", datetime.date.today().isoformat())
cov_row(8,"QA environment", "https://16.170.11.41/  (QA stack; data is durable — do not expect a clean DB)")
cov_row(9,"Login", "Use a NON-root user with the relevant role for permission cases (root bypasses all permission checks and will mask RBAC defects). Switch branch via the top-bar branch picker where a test needs a specific branch.")
cov_row(11,"How to run a case","1) Read Preconditions and set them up.  2) Follow Steps in the app.  3) Compare what you see to Expected Result.  4) Put Pass/Fail/Blocked in Status; if Fail, write what happened in Actual Result, set Severity, and log it on the 'Defects Log' tab.")
cov_row(13,"Status values","Not Run (default) · Pass · Fail · Blocked · N/A")
cov_row(14,"Severity (for fails)","Critical (blocks a core flow / data loss) · High · Medium · Low (cosmetic)")
cov_row(15,"Type","Positive (happy path) · Negative (rejection/validation) · Permission (RBAC) · Edge")
cov_row(16,"Priority","High = must pass for sign-off · Medium · Low")
cov_row(18,"Column guide","ID = stable case id · Area = feature · Route = where in the app · Permission = the RBAC code involved. 'Expected Result' is the pass criterion.")
cov_row(19,"Note","Some cases intentionally document a KNOWN issue (look for 'DEFECT'/'known issue' in the scenario or expected text) — verify the stated current behaviour.")
for r in range(4,20):
    cov.cell(r,2).alignment = TOP

# ============ SUMMARY (built after sheets; placeholder index) ============
summ = wb.create_sheet("Summary")

# ============ MODULE SHEETS ============
status_choices = '"Not Run,Pass,Fail,Blocked,N/A"'
sev_choices = '"-,Critical,High,Medium,Low"'
pri_choices = '"High,Medium,Low"'
type_choices = '"Positive,Negative,Permission,Edge"'

for mod in ordered_mods:
    sn = safe(mod); sheetname_of[mod] = sn
    ws = wb.create_sheet(sn)
    ws.sheet_view.showGridLines = False
    # title
    ncol = len(COLS)
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncol)
    tc = ws.cell(1,1,f"{mod}  —  UAT"); tc.font=Font(size=13,bold=True,color="FFFFFF"); tc.fill=TITLE_FILL; tc.alignment=Alignment(vertical="center")
    ws.row_dimensions[1].height = 22
    # header
    for ci,(name,w) in enumerate(COLS, start=1):
        c = ws.cell(2,ci,name); c.font=WHITE; c.fill=HEAD_FILL; c.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center"); c.border=BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    # data
    rws = mods[mod]
    for ri,row in enumerate(rws, start=3):
        vals = [row.get("id",""),row.get("area",""),row.get("title",""),row.get("type",""),row.get("priority",""),
                row.get("preconditions",""),row.get("steps",""),row.get("testData",""),row.get("expected",""),
                row.get("route",""),row.get("permission",""),"","Not Run","","","",""]
        for ci,v in enumerate(vals, start=1):
            c = ws.cell(ri,ci,v); c.border=BORDER
            c.alignment = CTR if ci in (4,5,13,14) else WRAP
            if ci==1: c.font=Font(bold=True,size=9)
        if ri % 2 == 1:
            for ci in range(1,ncol+1):
                if ws.cell(ri,ci).fill.fgColor.rgb in (None,"00000000"):
                    ws.cell(ri,ci).fill = PatternFill("solid", fgColor=GREY)
    last = len(rws)+2
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncol)}{last}"
    # data validations
    for choices, col in ((type_choices,4),(pri_choices,5),(status_choices,13),(sev_choices,14)):
        dv = DataValidation(type="list", formula1=choices, allow_blank=True)
        ws.add_data_validation(dv); dv.add(f"{get_column_letter(col)}3:{get_column_letter(col)}{last}")

# ============ fill SUMMARY (formulas referencing each module sheet) ============
summ.sheet_view.showGridLines = False
summ.merge_cells("A1:H1")
s=summ.cell(1,1,"UAT Summary Dashboard"); s.font=Font(size=15,bold=True,color="FFFFFF"); s.fill=TITLE_FILL; s.alignment=Alignment(horizontal="center",vertical="center")
summ.row_dimensions[1].height=26
hdr=["Module","Cases","Not Run","Pass","Fail","Blocked","N/A","Pass %"]
widths=[34,8,9,8,8,9,7,9]
for ci,(h,w) in enumerate(zip(hdr,widths),start=1):
    c=summ.cell(2,ci,h); c.font=WHITE; c.fill=HEAD_FILL; c.alignment=Alignment(horizontal="center"); c.border=BORDER
    summ.column_dimensions[get_column_letter(ci)].width=w
r=3
for mod in ordered_mods:
    sn=sheetname_of[mod]; q=f"'{sn}'"; rng=f"{q}!$M$3:$M$100000"
    summ.cell(r,1,mod).border=BORDER
    summ.cell(r,2,len(mods[mod])).border=BORDER
    summ.cell(r,3,f'=COUNTIF({rng},"Not Run")').border=BORDER
    summ.cell(r,4,f'=COUNTIF({rng},"Pass")').border=BORDER
    summ.cell(r,5,f'=COUNTIF({rng},"Fail")').border=BORDER
    summ.cell(r,6,f'=COUNTIF({rng},"Blocked")').border=BORDER
    summ.cell(r,7,f'=COUNTIF({rng},"N/A")').border=BORDER
    summ.cell(r,8,f'=IFERROR(D{r}/(D{r}+E{r}),"")').border=BORDER
    summ.cell(r,8).number_format="0%"
    for ci in range(2,9): summ.cell(r,ci).alignment=CTR
    r+=1
# totals
tr=r
summ.cell(tr,1,"TOTAL").font=BOLD
summ.cell(tr,2,f"=SUM(B3:B{r-1})").font=BOLD
for ci,L in zip(range(3,8),"CDEFG"):
    summ.cell(tr,ci,f"=SUM({L}3:{L}{r-1})").font=BOLD
summ.cell(tr,8,f'=IFERROR(D{tr}/(D{tr}+E{tr}),"")').font=BOLD; summ.cell(tr,8).number_format="0%"
for ci in range(1,9):
    summ.cell(tr,ci).fill=SUB_FILL; summ.cell(tr,ci).border=BORDER
    if ci>1: summ.cell(tr,ci).alignment=CTR
summ.freeze_panes="A3"

import os
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("WROTE", OUT)
print("sheets:", len(wb.sheetnames), "| cases:", len(rows))
